# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: http://doc.scrapy.org/en/latest/topics/item-pipeline.html
import codecs
import json
import requests
from czbbsspider import settings
import os

from czbbsspider.items import HeimaKbdlItem, HeimaKbdlDetailItem
import czbbsspider.utils as util_re
from openpyxl import Workbook, load_workbook

import datetime
import re


class FilterWordsPipeline(object):

    def __init__(self):
        # self.file = open("data.json", "wb")
        self.file = codecs.open(
            "scraped_data_utf8.json", "wb", encoding="utf-8")

    def process_item(self, item, spider):
        if item:
            line = json.dumps(dict(item), ensure_ascii=False) + "\n"
            self.file.write(line)
            return item

    def close_spider(self, spider):
        self.file.close()


# 过滤正文信息
class FilterTitlePipeLine(object):

    def process_item(self, item, spider):
        if isinstance(item, HeimaKbdlItem):
            if item['title']:
                return item
            else:
                # raise Exception(u'不是正文信息，自动忽略')
                pass
        else:
            return item


data_dic = {}

# 匹配账号信息


class ReadAccountPipeline(object):

    def __init__(self):
        data_dic.clear()
        workbook_ = load_workbook(filename=u'./czbbsspider/assets/2017.xlsx')
        sheetnames = workbook_.get_sheet_names()  # 获得表单名字
        self.ws = workbook_.get_sheet_by_name(sheetnames[0])  # 从工作表中提取某一表单

        # 把数据存到字典中 d = {“论坛名字”:"真实名字"}
        for rx in range(5, self.ws.max_row + 1):
            w2 = self.ws.cell(row=rx, column=4).value
            w4 = self.ws.cell(row=rx, column=5).value
            data_dic[w4] = w2
        print "********"+str(data_dic)

# 过滤当月时间
# 1.过滤列表 正常格式 2017-3-1
# 2.过滤回复
class FilterDataExcelPipeLine(object):

    def __init__(self):
        self._year = str(datetime.datetime.now().year)
        self._month = str(datetime.datetime.now().month)
        print u'\n\n初始化当前时间'+str(self._year)+'-'+str(self._month)+'\n\n'

    def process_item(self, item, spider):
        if isinstance(item, HeimaKbdlItem):
            updateTime = item['updateTime']
            print "------------------>"+updateTime.encode("utf-8")
            if " " in updateTime: 
                updateTime = updateTime[1:]
            ups = updateTime.split('-')
            #print ">>>>>>"+str(ups)
            # result = re.findall(util_re.time_re, updateTime)
            # print "????????"+str(result)
            # print ">>>>>>>>>>>>>FilterDataExcelPipeLine<<<<<"+str(updateTime)
            if ups:
                if self._year in ups[0] and self._month in ups[1]:
                    return item
                else:
                    pass
            else:
                return item
        elif isinstance(item, HeimaKbdlDetailItem):
            topstick = item['topstick']
            # replyTime = topstick['replyTime'] #发表于 7 天前
            replyTime = ''
            print '\n\n'+topstick
            if u'发布于' in topstick:
                topstick = topstick.replace("\n","")
                rs = topstick.split(' ')
                if rs and len(rs) > 2:
                    for i in rs:
                        if i and self._year in i:
                            replyTime = i
                    if replyTime:
                        print "******"+replyTime
                        result = replyTime.split("-")
                        if result and len(result) > 0:
                            if self._year == result[0] and self._month == result[1]:
                                return item 
                            else:
                                pass
                        else:
                            return item
                    else:
                        pass
            else:
                return item

# 讲数据统计到文档中

# 統計一篇文章記錄用戶的統計次數
article_count_times = {}
count_times_excel_d = {}  # 名：行


class WriteCleanDataAndCountTimes(object):

    def __init__(self):
        self.workbook_ = load_workbook(filename=u'./czbbsspider/assets/2017.xlsx')
        sheetnames = self.workbook_.get_sheet_names()  # 获得表单名字
        self.ws = self.workbook_.get_sheet_by_name(
            sheetnames[0])  # 从工作表中提取某一表单
        for rx in range(5, self.ws.max_row + 1):
            w5 = self.ws.cell(row=rx, column=4).value
            count_times_excel_d[w5] = rx
        for k, v in data_dic.items():
            article_count_times[v] = 0

# 記錄規則，一篇文章一人多次回復視為一次
    def process_item(self, item, spider):
        if isinstance(item, HeimaKbdlDetailItem):
            # 判断是否是要统计的用户
            topsticks = item['topsticks']
            print '*' * 50
            u_list = []
            for topstick in topsticks[1:]:
                # 获取论坛名字
                username_bbs = topstick['username']
                if not username_bbs:
                    continue
                if data_dic.has_key(username_bbs):
                    username = data_dic[username_bbs]
                    if username in u_list:
                        continue
                    u_list.append(username)
                    print u'检查到有  ' + username + u' 回复了当前文章'
                    # 只要發現這篇文章中有記錄用戶評論，那麼久終止 並且自動+ 1
                    article_count_times[
                        username] = article_count_times[username] + 1
            print '*' * 50
        return item

    def close_spider(self, spider):
        print u'\n\n统计数据如下 :'
        for k, v in article_count_times.items():
            if count_times_excel_d.has_key(k):
                print k, v
                self.ws.cell(row=count_times_excel_d[k], column=22).value = v
                self.workbook_.save(filename='./czbbsspider/assets/2017.xlsx')
        # self.ws.close()
# 将干净的数据写到xlsx中


class WriteCleanDataExcelPipeline(object):

    def __init__(self):
        print "------------------------------------------------"
        self.wb = Workbook()
        self.ws = self.wb.active
        th_line = ['板块', '姓名', '论坛账号', '发布时间', '序号', '文章名称', '链接', '回复/查看']
        self.ws.append(th_line)

    def process_item(self, item, spider):
        if isinstance(item, HeimaKbdlItem):
            bankuai_name = item['name']
            username_bbs = item['author']
            username = u''
            if data_dic.has_key(username_bbs):
                username = data_dic[username_bbs]
            # update_time = item['topsticks'][0].get('replyTime')
            update_time = item['updateTime']
            title = item['title']
            copy_url = item['href']
            reply_info = item['replyNum'] + "/" + item['sawNum']
            line = [bankuai_name, username, username_bbs, update_time,
                    '', title, copy_url, reply_info]  # 把数据中每一项整理出来
            self.ws.append(line)  # 将数据以行的形式添加到xlsx中
            today = datetime.date.today()
            self.wb.save('./czbbsspider/assets/heima-bankuai-list-' +
                         str(today) + '.xlsx')  # 保存xlsx文件
        return item
